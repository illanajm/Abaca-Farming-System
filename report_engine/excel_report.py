from io import BytesIO

from openpyxl import Workbook



def generate_excel_report(
        analysis,
        filtered_farmers,
        filtered_farms,
        filtered_pests
):


    output = BytesIO()

    wb = Workbook()



    ws = wb.active

    ws.title="AI Analysis"



    ws.append(
        ["Section","Result"]
    )


    for section in [
        "interpretation",
        "insights",
        "recommendations"
    ]:


        for item in analysis[section]:

            ws.append(
                [
                    section.upper(),
                    item
                ]
            )



    farm_sheet = wb.create_sheet(
        "Farm Data"
    )


    farm_sheet.append(
        filtered_farms.columns.tolist()
    )


    for row in filtered_farms.values.tolist():

        farm_sheet.append(row)



    farmer_sheet = wb.create_sheet(
        "Farmers"
    )


    farmer_sheet.append(
        filtered_farmers.columns.tolist()
    )


    for row in filtered_farmers.values.tolist():

        farmer_sheet.append(row)



    wb.save(output)


    output.seek(0)

    return output